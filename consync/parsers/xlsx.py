"""Excel (.xlsx) parser — reads constants from a spreadsheet.

Supports two layouts:

1. **Flat layout** (default for c_header, csv, etc.):
   Row 1: Header row (Name, Value, Unit, Description)
   Row 2+: Data rows

2. **Table layout** (for c_struct_table):
   Row 1: Motor Variant | field1 | field2 | field3 | ...
   Row 2+: variant_name | val1   | val2   | val3   | ...
   - Multiple sheets = multiple variants
   - Auto-detected when first column header matches "Motor Variant"
   - Reconstructs full metadata for bidirectional sync with C files

Column mapping is flexible — auto-detects by header names.
"""

from __future__ import annotations

import re
from pathlib import Path

from consync.models import Constant
from consync.parsers import register


# Column header aliases (case-insensitive matching)
NAME_ALIASES = {"name", "constant", "parameter", "variable", "symbol", "id"}
VALUE_ALIASES = {"value", "val", "data", "number", "amount"}
UNIT_ALIASES = {"unit", "units", "uom", "dimension"}
DESC_ALIASES = {"description", "desc", "comment", "note", "notes", "info"}

# Table layout detection
TABLE_FIRST_COL_ALIASES = {"motor variant", "variant", "row", "label", "name"}


def _find_column(headers: list[str], aliases: set[str]) -> int | None:
    """Find column index matching any alias (case-insensitive)."""
    for i, h in enumerate(headers):
        if h and h.strip().lower() in aliases:
            return i
    return None


def _is_table_layout(ws) -> bool:
    """Detect if a worksheet uses table layout (first col = Motor Variant)."""
    first_header = ws.cell(1, 1).value
    if first_header and str(first_header).strip().lower() in TABLE_FIRST_COL_ALIASES:
        # Additional check: second column should NOT be "Value"
        second_header = ws.cell(1, 2).value
        if second_header and str(second_header).strip().lower() in VALUE_ALIASES:
            return False
        return True
    return False


def _sanitize_label(label: str) -> str:
    """Convert a row label to a valid C-style identifier prefix."""
    sanitized = re.sub(r"[^a-zA-Z0-9]+", "_", label.strip())
    return sanitized.strip("_")


@register("xlsx")
def parse_xlsx(filepath: str | Path, **kwargs) -> list[Constant]:
    """Parse constants from an Excel file.

    Auto-detects layout:
    - Table layout → reads all sheets, reconstructs row_label/field metadata
    - Flat layout → reads active sheet as Name/Value/Unit/Description

    Args:
        filepath: Path to .xlsx file.
        sheet: Sheet name or index (default: auto).

    Returns:
        List of Constant objects.
    """
    import openpyxl

    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Check if this is table layout by examining the first data sheet
    data_sheets = [s for s in wb.sheetnames if s.lower() != "info"]
    if data_sheets:
        first_ws = wb[data_sheets[0]]
        if _is_table_layout(first_ws):
            return _parse_table_layout(wb, data_sheets, **kwargs)

    # Fall back to flat layout
    return _parse_flat_layout(wb, **kwargs)


def _parse_table_layout(wb, data_sheets: list[str], **kwargs) -> list[Constant]:
    """Parse table-layout Excel (one sheet per variant, rows=motors, cols=fields)."""
    constants: list[Constant] = []

    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        variant = sheet_name  # Sheet name = variant name

        # Read field names from header row (skip first column = "Motor Variant")
        headers = []
        for col in range(2, ws.max_column + 1):
            h = ws.cell(1, col).value
            headers.append(str(h).strip() if h else f"field_{col - 2}")

        # Read data rows
        for row in range(2, ws.max_row + 1):
            row_label = ws.cell(row, 1).value
            if not row_label:
                continue
            row_label = str(row_label).strip()
            name_prefix = _sanitize_label(row_label)

            for col_idx, field_name in enumerate(headers):
                cell_value = ws.cell(row, col_idx + 2).value
                if cell_value is None:
                    continue

                const_name = f"{name_prefix}__{field_name}"

                # Determine value type
                if isinstance(cell_value, (int, float)):
                    value = cell_value
                    is_expression = False
                else:
                    str_val = str(cell_value).strip()
                    # Try numeric parse
                    try:
                        value = int(str_val)
                        is_expression = False
                    except ValueError:
                        try:
                            value = float(str_val)
                            is_expression = False
                        except ValueError:
                            value = str_val
                            is_expression = str_val.upper() not in ("TRUE", "FALSE")

                constants.append(Constant(
                    name=const_name,
                    value=value,
                    unit="",
                    description=f"Row: {row_label}, Field: {field_name}",
                    metadata={
                        "row_label": row_label,
                        "field": field_name,
                        "field_index": col_idx,
                        "variant": variant,
                        "is_expression": is_expression,
                    },
                ))

    return constants


def _parse_flat_layout(wb, **kwargs) -> list[Constant]:
    """Parse flat-layout Excel (Name/Value/Unit/Description columns)."""
    sheet = kwargs.get("sheet")
    if sheet is not None:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
    else:
        ws = wb.active

    # Read header row
    header_row = [str(cell.value or "").strip() for cell in ws[1]]

    # Auto-detect columns
    name_col = _find_column(header_row, NAME_ALIASES)
    value_col = _find_column(header_row, VALUE_ALIASES)
    unit_col = _find_column(header_row, UNIT_ALIASES)
    desc_col = _find_column(header_row, DESC_ALIASES)

    # Fallback to positional: A=Name, B=Value, C=Unit, D=Description
    if name_col is None:
        name_col = 0
    if value_col is None:
        value_col = 1
    if unit_col is None:
        unit_col = 2 if len(header_row) > 2 else None
    if desc_col is None:
        desc_col = 3 if len(header_row) > 3 else None

    constants: list[Constant] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[name_col] is None:
            continue

        name = str(row[name_col]).strip()
        if not name:
            continue

        raw_value = row[value_col] if value_col < len(row) else None
        if raw_value is None:
            continue

        # Preserve numeric types
        if isinstance(raw_value, (int, float)):
            value = raw_value
        else:
            # Try to parse as number
            try:
                value = int(str(raw_value))
            except ValueError:
                try:
                    value = float(str(raw_value))
                except ValueError:
                    value = str(raw_value)

        unit = ""
        if unit_col is not None and unit_col < len(row) and row[unit_col]:
            unit = str(row[unit_col]).strip()

        description = ""
        if desc_col is not None and desc_col < len(row) and row[desc_col]:
            description = str(row[desc_col]).strip()

        constants.append(Constant(name=name, value=value, unit=unit, description=description))

    return constants
