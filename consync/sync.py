"""Sync engine — the core logic that ties parsers, renderers, and state together."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any

from consync.backup import backup_file
from consync.config import load_config
from consync.lock import SyncLock, LockError
from consync.logging_config import write_audit_entry
from consync.models import ConsyncConfig, MappingConfig, SyncDirection
from consync.parsers import get_parser
from consync.renderers import get_renderer
from consync.state import SyncState, compute_hash
from consync.validators import parse_validators, validate_constants

logger = logging.getLogger(__name__)


class SyncResult(Enum):
    """Outcome of a sync operation."""
    SYNCED_SOURCE_TO_TARGET = "source → target"
    SYNCED_TARGET_TO_SOURCE = "target → source"
    ALREADY_IN_SYNC = "already in sync"
    CONFLICT = "conflict"
    SKIPPED = "skipped"
    ERROR = "error"


@dataclass
class SyncReport:
    """Result of syncing one mapping."""
    source: str
    target: str
    result: SyncResult
    count: int = 0  # number of constants synced
    message: str = ""


def sync(
    config_path: str | Path | None = None,
    force_direction: str | None = None,
    dry_run: bool = False,
) -> list[SyncReport]:
    """Run sync for all mappings in config.

    Args:
        config_path: Path to .consync.yaml (auto-find if None).
        force_direction: Override direction ("source" or "target").
        dry_run: If True, report what would happen without writing files.

    Returns:
        List of SyncReport for each mapping.
    """
    cfg = load_config(config_path)
    config_dir = _config_dir(config_path)
    state = SyncState(config_dir / cfg.state_file)

    reports: list[SyncReport] = []

    try:
        with SyncLock(config_dir):
            for mapping in cfg.mappings:
                report = _sync_one(mapping, config_dir, state, cfg, force_direction, dry_run)
                reports.append(report)
    except LockError as e:
        logger.error("Lock conflict: %s", e)
        reports.append(SyncReport(
            source="*", target="*",
            result=SyncResult.ERROR,
            message=str(e),
        ))

    return reports


def check(config_path: str | Path | None = None) -> list[SyncReport]:
    """Verify all mappings are in sync (CI mode).

    Returns reports — any with result != ALREADY_IN_SYNC means out-of-sync.
    Does NOT write any files.
    """
    cfg = load_config(config_path)
    config_dir = _config_dir(config_path)

    reports: list[SyncReport] = []
    for mapping in cfg.mappings:
        report = _check_one(mapping, config_dir)
        reports.append(report)

    return reports


def _config_dir(config_path: str | Path | None) -> Path:
    """Resolve the directory containing the config file."""
    if config_path is None:
        from consync.config import find_config
        found = find_config()
        if found:
            return found.parent
        return Path.cwd()
    return Path(config_path).parent


def _resolve_path(filepath: str, config_dir: Path) -> Path:
    """Resolve a path relative to config directory."""
    p = Path(filepath)
    if p.is_absolute():
        return p
    return (config_dir / p).resolve()


def _sync_one(
    mapping: MappingConfig,
    config_dir: Path,
    state: SyncState,
    cfg: ConsyncConfig,
    force_direction: str | None,
    dry_run: bool,
) -> SyncReport:
    """Sync a single mapping."""
    source_path = _resolve_path(mapping.source, config_dir)
    target_path = _resolve_path(mapping.target, config_dir)
    key = state.mapping_key(mapping.source, mapping.target)

    try:
        # Determine direction
        direction = _determine_direction(
            mapping, source_path, target_path, state, key, cfg.on_conflict, force_direction
        )

        if direction is None:
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.ALREADY_IN_SYNC,
                message="Files are already in sync.",
            )

        if direction == "conflict":
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.CONFLICT,
                message="Both files changed. Use --from source or --from target to resolve.",
            )

        if dry_run:
            result = (SyncResult.SYNCED_SOURCE_TO_TARGET
                      if direction == "source" else SyncResult.SYNCED_TARGET_TO_SOURCE)
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=result,
                message=f"[DRY RUN] Would sync {result.value}",
            )

        # Execute sync (with backup before overwrite)
        if direction == "source":
            constants = _parse_file(source_path, mapping.source_format, mapping.parser_options)
            # Validate before writing
            validation = _validate_if_needed(constants, mapping)
            if validation and not validation.ok:
                return SyncReport(
                    source=mapping.source, target=mapping.target,
                    result=SyncResult.ERROR,
                    message=f"Validation failed: {validation.errors[0].message}"
                    + (f" (+{len(validation.errors)-1} more)" if len(validation.errors) > 1 else ""),
                )
            backup_file(target_path, backup_dir=config_dir / ".consync" / "backups")
            _render_file(constants, target_path, mapping.target_format, mapping)
            result = SyncResult.SYNCED_SOURCE_TO_TARGET
            logger.info(
                "Synced %s → %s (%d constants)",
                mapping.source, mapping.target, len(constants),
            )
        else:
            constants = _parse_file(target_path, mapping.target_format, mapping.parser_options)
            # Validate before writing
            validation = _validate_if_needed(constants, mapping)
            if validation and not validation.ok:
                return SyncReport(
                    source=mapping.source, target=mapping.target,
                    result=SyncResult.ERROR,
                    message=f"Validation failed: {validation.errors[0].message}"
                    + (f" (+{len(validation.errors)-1} more)" if len(validation.errors) > 1 else ""),
                )
            backup_file(source_path, backup_dir=config_dir / ".consync" / "backups")
            _render_file(constants, source_path, mapping.source_format, mapping)
            result = SyncResult.SYNCED_TARGET_TO_SOURCE
            logger.info(
                "Synced %s → %s (%d constants)",
                mapping.target, mapping.source, len(constants),
            )

        # Log individual constant values at DEBUG level
        for c in constants:
            logger.debug("  %s = %r%s", c.name, c.value, f" ({c.unit})" if c.unit else "")

        # Write structured audit entry
        write_audit_entry(
            direction=result.value,
            source=mapping.source,
            target=mapping.target,
            constants=constants,
            result="synced",
            dry_run=dry_run,
            audit_file=config_dir / ".consync.audit.jsonl",
        )

        # Update state with hashes of BOTH files after sync
        src_constants = _parse_file(source_path, mapping.source_format, mapping.parser_options)
        src_hash = compute_hash(src_constants)
        # For target formats without a parser, use source hash (they're equivalent after sync)
        try:
            tgt_constants = _parse_file(target_path, mapping.target_format, mapping.parser_options)
            tgt_hash = compute_hash(tgt_constants)
        except (ValueError, FileNotFoundError):
            tgt_hash = src_hash
        state.set_hash(key, src_hash, tgt_hash)

        return SyncReport(
            source=mapping.source, target=mapping.target,
            result=result,
            count=len(constants),
            message=f"{len(constants)} constants synced ({result.value})",
        )

    except Exception as e:
        logger.error("Sync failed for %s ↔ %s: %s", mapping.source, mapping.target, e)
        return SyncReport(
            source=mapping.source, target=mapping.target,
            result=SyncResult.ERROR,
            message=str(e),
        )


def _check_one(mapping: MappingConfig, config_dir: Path) -> SyncReport:
    """Check if a single mapping is in sync."""
    source_path = _resolve_path(mapping.source, config_dir)
    target_path = _resolve_path(mapping.target, config_dir)

    try:
        if not source_path.exists():
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.ERROR,
                message=f"Source file not found: {source_path}",
            )
        if not target_path.exists():
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.ERROR,
                message=f"Target file not found: {target_path}",
            )

        src_constants = _parse_file(source_path, mapping.source_format, mapping.parser_options)
        tgt_constants = _parse_file(target_path, mapping.target_format, mapping.parser_options)

        src_hash = compute_hash(src_constants)
        tgt_hash = compute_hash(tgt_constants)

        if src_hash == tgt_hash:
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.ALREADY_IN_SYNC,
                count=len(src_constants),
                message=f"In sync ({len(src_constants)} constants).",
            )
        else:
            return SyncReport(
                source=mapping.source, target=mapping.target,
                result=SyncResult.CONFLICT,
                count=len(src_constants),
                message=f"OUT OF SYNC. Source has {len(src_constants)} constants, "
                        f"target has {len(tgt_constants)}. Run 'consync sync' to fix.",
            )

    except Exception as e:
        return SyncReport(
            source=mapping.source, target=mapping.target,
            result=SyncResult.ERROR,
            message=str(e),
        )


def _determine_direction(
    mapping: MappingConfig,
    source_path: Path,
    target_path: Path,
    state: SyncState,
    key: str,
    on_conflict: str,
    force_direction: str | None,
) -> str | None:
    """Determine sync direction based on state hashes and config.

    Returns:
        "source" — sync source → target
        "target" — sync target → source
        "conflict" — both changed, can't auto-resolve
        None — already in sync
    """
    # Forced direction overrides everything
    if force_direction:
        return "source" if force_direction.lower() in ("source", "xlsx", "s") else "target"

    # One-way modes: always sync in configured direction
    if mapping.direction == SyncDirection.SOURCE_TO_TARGET:
        if not target_path.exists():
            return "source"
        # Check if source changed since last sync (use state hash)
        src_constants = _parse_file(source_path, mapping.source_format, mapping.parser_options)
        src_hash = compute_hash(src_constants)
        last_src_hash = state.get_hash(key, "source") if state else None
        if last_src_hash and src_hash == last_src_hash:
            return None  # Source hasn't changed
        return "source"

    if mapping.direction == SyncDirection.TARGET_TO_SOURCE:
        if not source_path.exists():
            return "target"
        # Check if target changed since last sync
        try:
            tgt_constants = _parse_file(target_path, mapping.target_format, mapping.parser_options)
            tgt_hash = compute_hash(tgt_constants)
            last_tgt_hash = state.get_hash(key, "target") if state else None
            if last_tgt_hash and tgt_hash == last_tgt_hash:
                return None
        except (ValueError, FileNotFoundError):
            pass
        return "target"

    # Bidirectional: use state hashes to detect which side changed
    if not source_path.exists():
        return "target"
    if not target_path.exists():
        return "source"

    src_constants = _parse_file(source_path, mapping.source_format, mapping.parser_options)
    tgt_constants = _parse_file(target_path, mapping.target_format, mapping.parser_options)
    cur_src = compute_hash(src_constants)
    cur_tgt = compute_hash(tgt_constants)

    prev_src = state.get_hash(key, "source")
    prev_tgt = state.get_hash(key, "target")

    # No prior state — treat source as truth
    if prev_src is None or prev_tgt is None:
        if cur_src == cur_tgt:
            return None
        return "source"

    src_changed = cur_src != prev_src
    tgt_changed = cur_tgt != prev_tgt

    if not src_changed and not tgt_changed:
        return None
    if src_changed and not tgt_changed:
        return "source"
    if tgt_changed and not src_changed:
        return "target"

    # Both changed — conflict
    if on_conflict == "source_wins":
        return "source"
    elif on_conflict == "target_wins":
        return "target"
    else:
        return "conflict"


def _parse_file(filepath: Path, format_name: str, parser_options: dict | None = None) -> list:
    """Parse a file using the appropriate parser."""
    parser = get_parser(format_name)
    opts = parser_options or {}
    return parser(filepath, **opts)


def _render_file(constants: list, filepath: Path, format_name: str, mapping: MappingConfig):
    """Render constants to a file using the appropriate renderer."""
    # Special case: xlsx output needs openpyxl writer (not a simple renderer)
    if format_name == "xlsx":
        _write_xlsx(constants, filepath, mapping)
        return

    renderer = get_renderer(format_name)
    renderer(constants, filepath, config=mapping)


def _validate_if_needed(constants: list, mapping: MappingConfig):
    """Run validation hooks if configured. Returns ValidationResult or None."""
    if not mapping.validators:
        return None
    rules = parse_validators(mapping.validators)
    result = validate_constants(constants, rules)
    if not result.ok:
        for err in result.errors:
            logger.warning("Validation error: %s", err.message)
    return result


def _write_xlsx(constants: list, filepath: Path, mapping: MappingConfig):
    """Write constants back to an Excel file.

    If constants come from a table parser (have row_label + field metadata),
    writes in TABLE layout: rows = variants, columns = fields.
    Otherwise falls back to the flat Name/Value/Unit/Description layout.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    filepath.parent.mkdir(parents=True, exist_ok=True)

    # Detect if data came from a table parser (c_struct_table)
    is_table = (
        constants
        and constants[0].metadata.get("row_label") is not None
        and constants[0].metadata.get("field") is not None
    )

    if is_table:
        _write_xlsx_table(constants, filepath, mapping)
    else:
        _write_xlsx_flat(constants, filepath, mapping)


def _write_xlsx_table(constants: list, filepath: Path, mapping: MappingConfig):
    """Write constants as a proper table: rows = motor variants, columns = fields.

    If constants have different metadata['variant'] values, creates one sheet per variant.
    Otherwise creates a single sheet.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict

    # Group constants by variant
    variants_map: OrderedDict[str, list] = OrderedDict()
    for c in constants:
        v = c.metadata.get("variant", "_default")
        if v not in variants_map:
            variants_map[v] = []
        variants_map[v].append(c)

    wb = openpyxl.Workbook()
    # Remove default sheet — we'll create named ones
    wb.remove(wb.active)

    for variant_name, variant_constants in variants_map.items():
        sheet_name = variant_name if variant_name != "_default" else "Parameters"
        # Excel sheet name max 31 chars
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_table_sheet(ws, variant_constants, mapping)

    # Add Info sheet
    ws_info = wb.create_sheet("Info")
    variant_cfg = mapping.parser_options.get("variant", "")
    table_var = mapping.parser_options.get("table_var", "(auto-detected)")
    ws_info.append(["Source File", mapping.source])
    ws_info.append(["Struct Array", table_var])
    ws_info.append(["Variants", ", ".join(variants_map.keys())])
    ws_info.append(["Total Constants", len(constants)])
    ws_info.append(["Generated By", "consync"])
    ws_info.column_dimensions["A"].width = 16
    ws_info.column_dimensions["B"].width = 50

    wb.save(filepath)


def _write_table_sheet(ws, constants: list, mapping: MappingConfig):
    """Write one variant's constants as a table in a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import OrderedDict
    from typing import Any

    # Group constants by row_label, preserving order
    rows: OrderedDict[str, dict[str, Any]] = OrderedDict()
    fields_ordered: list[str] = []

    for c in constants:
        row_label = c.metadata["row_label"]
        field_name = c.metadata["field"]
        if row_label not in rows:
            rows[row_label] = {}
        rows[row_label][field_name] = c
        if field_name not in fields_ordered:
            fields_ordered.append(field_name)

    # --- Styling ---
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    label_font = Font(bold=True, size=10)
    value_font = Font(name="Courier New", size=10)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill = PatternFill("solid", fgColor="D6E4F0")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    expr_font = Font(name="Courier New", size=9, italic=True, color="666666")

    # --- Header row ---
    ws.cell(row=1, column=1, value="Motor Variant")
    for col_idx, field_name in enumerate(fields_ordered, 2):
        ws.cell(row=1, column=col_idx, value=field_name)

    # Style header
    for col in range(1, len(fields_ordered) + 2):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # --- Data rows ---
    for row_idx, (row_label, field_map) in enumerate(rows.items(), 2):
        # Motor variant name in column 1
        ws.cell(row=row_idx, column=1, value=row_label)
        ws.cell(row=row_idx, column=1).font = label_font
        ws.cell(row=row_idx, column=1).border = border
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        ws.cell(row=row_idx, column=1).fill = fill

        for col_idx, field_name in enumerate(fields_ordered, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right")

            c = field_map.get(field_name)
            if c is None:
                cell.value = ""
                continue

            is_expr = c.metadata.get("is_expression", False)
            if is_expr:
                # Show expression/macro as-is (string)
                cell.value = c.value
                cell.font = expr_font
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(c.value, float):
                cell.value = c.value
                cell.font = value_font
                # Use scientific notation for very small/large numbers
                if abs(c.value) < 0.001 or abs(c.value) >= 100000:
                    cell.number_format = "0.000E+00"
                else:
                    cell.number_format = "0.######"
            elif isinstance(c.value, int):
                cell.value = c.value
                cell.font = value_font
                cell.number_format = "0"
            else:
                cell.value = str(c.value)
                cell.font = value_font

    # --- Column widths ---
    ws.column_dimensions["A"].width = 18  # Motor Variant
    for col_idx in range(2, len(fields_ordered) + 2):
        col_letter = get_column_letter(col_idx)
        field_name = fields_ordered[col_idx - 2]
        # Width based on field name length, min 10
        ws.column_dimensions[col_letter].width = max(len(field_name) + 3, 10)

    # Freeze first row + first column
    ws.freeze_panes = "B2"


def _write_xlsx_flat(constants: list, filepath: Path, mapping: MappingConfig):
    """Write constants in flat Name/Value/Unit/Description layout (original behavior)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    filepath.parent.mkdir(parents=True, exist_ok=True)

    if filepath.exists():
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Constants"
        headers = ["Name", "Value", "Unit", "Description"]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="CCCCCC")
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    even_fill = PatternFill("solid", fgColor="D6E4F0")
    value_font = Font(name="Courier New", size=10)

    for i, c in enumerate(constants, 2):
        ws.cell(row=i, column=1).value = c.name
        ws.cell(row=i, column=2).value = c.value
        ws.cell(row=i, column=3).value = c.unit
        ws.cell(row=i, column=4).value = c.description
        fill = even_fill if i % 2 == 0 else None
        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            if fill:
                cell.fill = fill
            cell.border = border
            if col == 2:
                cell.font = value_font
                cell.number_format = "0.00000000000000"
                cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 36
    ws.freeze_panes = "A2"

    wb.save(filepath)
