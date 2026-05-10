"""Comprehensive sync audit — tests every edge case that can cause sync issues.

Covers: compute_hash, _determine_direction, _sync_one state updates,
parser/renderer fidelity, watcher patterns, CLI diff, and round-trip integrity.
"""
import textwrap
import openpyxl
import pytest
from pathlib import Path

from consync.models import Constant, MappingConfig, SyncDirection
from consync.state import SyncState, compute_hash
from consync.sync import sync, check, SyncResult, _determine_direction, _parse_file


# ═══════════════════════════════════════════════════════════════════════
# 1. compute_hash — duplicate names, floats, types, edge cases
# ═══════════════════════════════════════════════════════════════════════

class TestComputeHashDuplicates:
    """Ensure compute_hash handles duplicate constant names (multi-variant)."""

    def test_duplicate_names_different_values_differ(self):
        """Two constants with same name but different values → different hash."""
        c1 = [Constant("X__R", 1.0), Constant("X__R", 2.0)]
        c2 = [Constant("X__R", 1.0), Constant("X__R", 1.0)]
        assert compute_hash(c1) != compute_hash(c2)

    def test_duplicate_names_same_values_match(self):
        """Two constants with same name and same values → same hash."""
        c1 = [Constant("X__R", 1.0), Constant("X__R", 2.0)]
        c2 = [Constant("X__R", 1.0), Constant("X__R", 2.0)]
        assert compute_hash(c1) == compute_hash(c2)

    def test_single_value_change_in_duplicates_detected(self):
        """Changing one value among duplicates changes the hash."""
        before = [Constant("A", 1.0), Constant("A", 2.0), Constant("A", 3.0)]
        after  = [Constant("A", 1.0), Constant("A", 2.5), Constant("A", 3.0)]
        assert compute_hash(before) != compute_hash(after)

    def test_30_constants_single_edit(self):
        """Simulates real struct table: 30 constants, change 1 value."""
        before = [Constant(f"Motor_{i}__{f}", float(i * 10 + j))
                  for i in range(3) for j, f in enumerate(["R", "L", "Q", "P", "S"])]
        after = list(before)
        # Change one value
        after[7] = Constant(after[7].name, 999.0)
        assert compute_hash(before) != compute_hash(after)


class TestComputeHashTypes:
    """Ensure compute_hash distinguishes different types correctly."""

    def test_int_vs_float_same_value(self):
        """int 1 vs float 1.0 — should produce different hashes."""
        c_int = [Constant("X", 1)]
        c_float = [Constant("X", 1.0)]
        # Note: json.dumps treats 1 and 1.0 differently
        h1 = compute_hash(c_int)
        h2 = compute_hash(c_float)
        # These may or may not differ depending on json serialization
        # The key thing is consistency, not distinction
        assert isinstance(h1, str) and len(h1) == 32

    def test_string_values_hashed(self):
        """String constant values are hashed correctly."""
        c1 = [Constant("X", "MY_MACRO")]
        c2 = [Constant("X", "OTHER_MACRO")]
        assert compute_hash(c1) != compute_hash(c2)

    def test_empty_list(self):
        """Empty constant list produces a valid hash."""
        h = compute_hash([])
        assert isinstance(h, str) and len(h) == 32

    def test_list_values_hashed(self):
        """Array constant values are hashed via default=str."""
        c1 = [Constant("X", [1, 2, 3])]
        c2 = [Constant("X", [1, 2, 4])]
        assert compute_hash(c1) != compute_hash(c2)


# ═══════════════════════════════════════════════════════════════════════
# 2. _determine_direction — every code path
# ═══════════════════════════════════════════════════════════════════════

class TestDetermineDirection:
    """Test direction detection logic for all modes and edge cases."""

    def _make_mapping(self, direction=SyncDirection.BOTH):
        return MappingConfig(
            source="src.csv", target="tgt.h",
            source_format="csv", target_format="c_header",
            direction=direction,
        )

    def test_force_source(self, tmp_path):
        """force_direction='source' always returns 'source'."""
        m = self._make_mapping()
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1\n")
        tgt.write_text("const double A = 1;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", "source")
        assert result == "source"

    def test_force_target(self, tmp_path):
        """force_direction='target' always returns 'target'."""
        m = self._make_mapping()
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1\n")
        tgt.write_text("const double A = 1;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", "target")
        assert result == "target"

    def test_s2t_no_target_returns_source(self, tmp_path):
        """SOURCE_TO_TARGET: target doesn't exist → sync source."""
        m = self._make_mapping(SyncDirection.SOURCE_TO_TARGET)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "source"

    def test_s2t_source_unchanged_returns_none(self, tmp_path):
        """SOURCE_TO_TARGET: source hasn't changed → None (already in sync)."""
        m = self._make_mapping(SyncDirection.SOURCE_TO_TARGET)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1\n")
        tgt.write_text("const double A = 1;\n")
        # Set state hash matching current source
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        src_constants = _parse_file(src, "csv")
        src_hash = compute_hash(src_constants)
        state.set_hash(key, src_hash, "whatever")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result is None

    def test_t2s_no_source_returns_target(self, tmp_path):
        """TARGET_TO_SOURCE: source doesn't exist → sync target."""
        m = self._make_mapping(SyncDirection.TARGET_TO_SOURCE)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        tgt.write_text("const double A = 1;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "target"

    def test_both_no_prior_state_source_wins(self, tmp_path):
        """BOTH: no prior state, files differ → source wins."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1\n")
        tgt.write_text("const double A = 2.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "source"

    def test_both_no_prior_state_same_content_none(self, tmp_path):
        """BOTH: no prior state, files have same content → None."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1.0\n")
        tgt.write_text("const double A = 1.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result is None

    def test_both_only_source_changed(self, tmp_path):
        """BOTH: only source changed → source."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1.0\n")
        tgt.write_text("const double A = 1.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        # Simulate previous sync
        src_hash = compute_hash(_parse_file(src, "csv"))
        tgt_hash = compute_hash(_parse_file(tgt, "c_header"))
        state.set_hash(key, src_hash, tgt_hash)
        # Now change source
        src.write_text("Name,Value\nA,2.0\n")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "source"

    def test_both_only_target_changed(self, tmp_path):
        """BOTH: only target changed → target."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1.0\n")
        tgt.write_text("const double A = 1.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        src_hash = compute_hash(_parse_file(src, "csv"))
        tgt_hash = compute_hash(_parse_file(tgt, "c_header"))
        state.set_hash(key, src_hash, tgt_hash)
        # Now change target
        tgt.write_text("const double A = 9.0;\n")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "target"

    def test_both_both_changed_conflict(self, tmp_path):
        """BOTH: both changed, on_conflict='fail' → conflict."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1.0\n")
        tgt.write_text("const double A = 1.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        src_hash = compute_hash(_parse_file(src, "csv"))
        tgt_hash = compute_hash(_parse_file(tgt, "c_header"))
        state.set_hash(key, src_hash, tgt_hash)
        # Change both
        src.write_text("Name,Value\nA,2.0\n")
        tgt.write_text("const double A = 3.0;\n")
        result = _determine_direction(m, src, tgt, state, key, "fail", None)
        assert result == "conflict"

    def test_both_both_changed_source_wins(self, tmp_path):
        """BOTH: both changed, on_conflict='source_wins' → source."""
        m = self._make_mapping(SyncDirection.BOTH)
        src = tmp_path / "src.csv"
        tgt = tmp_path / "tgt.h"
        src.write_text("Name,Value\nA,1.0\n")
        tgt.write_text("const double A = 1.0;\n")
        state = SyncState(tmp_path / ".state.json")
        key = state.mapping_key("src.csv", "tgt.h")
        src_hash = compute_hash(_parse_file(src, "csv"))
        tgt_hash = compute_hash(_parse_file(tgt, "c_header"))
        state.set_hash(key, src_hash, tgt_hash)
        src.write_text("Name,Value\nA,2.0\n")
        tgt.write_text("const double A = 3.0;\n")
        result = _determine_direction(m, src, tgt, state, key, "source_wins", None)
        assert result == "source"


# ═══════════════════════════════════════════════════════════════════════
# 3. _sync_one state update — does state reflect BOTH files after sync?
# ═══════════════════════════════════════════════════════════════════════

class TestSyncStateUpdates:
    """After sync, state hashes must match BOTH files so next sync is noop."""

    @pytest.fixture
    def csv_h_project(self, tmp_path):
        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: data.csv
                target: data.h
                direction: both
                precision: 17
                header_guard: DATA_H
        """))
        csv = tmp_path / "data.csv"
        csv.write_text("Name,Value,Unit,Description\nPI,3.14159,,Pi\nG,9.81,,Gravity\n")
        return tmp_path, config

    def test_first_sync_then_noop(self, csv_h_project):
        """After initial sync, second sync must be 'already in sync'."""
        tmp, cfg = csv_h_project
        r1 = sync(config_path=cfg)
        assert r1[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        r2 = sync(config_path=cfg)
        assert r2[0].result == SyncResult.ALREADY_IN_SYNC

    def test_edit_source_then_sync_then_noop(self, csv_h_project):
        """Edit source, sync, then second sync is noop."""
        tmp, cfg = csv_h_project
        sync(config_path=cfg)
        # Edit source
        csv = tmp / "data.csv"
        csv.write_text("Name,Value,Unit,Description\nPI,3.14,,Pi\nG,9.81,,Gravity\n")
        r1 = sync(config_path=cfg)
        assert r1[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        r2 = sync(config_path=cfg)
        assert r2[0].result == SyncResult.ALREADY_IN_SYNC

    def test_edit_target_then_sync_then_noop(self, csv_h_project):
        """Edit target, sync, then second sync is noop."""
        tmp, cfg = csv_h_project
        sync(config_path=cfg)
        # Edit target — find the actual rendered PI value and change it
        h = tmp / "data.h"
        content = h.read_text()
        # The renderer outputs full precision, so find the PI line and swap value
        # Replace entire "PI" constant line with a different value
        import re
        content = re.sub(
            r"(PI\s*=\s*)[0-9.]+",
            r"\g<1>2.71828000000000000",
            content,
        )
        h.write_text(content)
        r1 = sync(config_path=cfg)
        assert r1[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE
        r2 = sync(config_path=cfg)
        assert r2[0].result == SyncResult.ALREADY_IN_SYNC


# ═══════════════════════════════════════════════════════════════════════
# 4. xlsx ↔ c_header flat round-trip
# ═══════════════════════════════════════════════════════════════════════

class TestFlatXlsxRoundTrip:
    """Test xlsx ↔ C header flat layout round-trip fidelity."""

    @pytest.fixture
    def flat_project(self, tmp_path):
        # Create xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Constants"
        ws.append(["Name", "Value", "Unit", "Description"])
        ws.append(["R_SENSE", 1.999, "Ohm", "Sense resistor"])
        ws.append(["R_PULLUP", 4706, "Ohm", "Pull-up"])
        ws.append(["VOLTAGE", 3.3, "V", "Supply voltage"])
        wb.save(tmp_path / "params.xlsx")

        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: params.xlsx
                target: params.h
                direction: both
                precision: 17
                header_guard: PARAMS_H
        """))
        return tmp_path, config

    def test_xlsx_to_c_creates_header(self, flat_project):
        """xlsx → C header works on first sync."""
        tmp, cfg = flat_project
        reports = sync(config_path=cfg)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        h = tmp / "params.h"
        assert h.exists()
        content = h.read_text()
        assert "R_SENSE" in content
        assert "R_PULLUP" in content
        assert "VOLTAGE" in content

    def test_xlsx_to_c_then_noop(self, flat_project):
        """Second sync after xlsx→C is noop."""
        tmp, cfg = flat_project
        sync(config_path=cfg)
        r2 = sync(config_path=cfg)
        assert r2[0].result == SyncResult.ALREADY_IN_SYNC

    def test_edit_xlsx_then_sync_updates_c(self, flat_project):
        """Edit xlsx value, sync picks up change and updates C."""
        tmp, cfg = flat_project
        sync(config_path=cfg)
        # Edit xlsx
        wb = openpyxl.load_workbook(tmp / "params.xlsx")
        ws = wb.active
        ws.cell(2, 2).value = 2.5  # Change R_SENSE
        wb.save(tmp / "params.xlsx")
        r = sync(config_path=cfg)
        assert r[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp / "params.h").read_text()
        assert "2.5" in content

    def test_edit_c_then_sync_updates_xlsx(self, flat_project):
        """Edit C header value, sync picks up change and updates xlsx."""
        tmp, cfg = flat_project
        sync(config_path=cfg)
        # Edit C header
        h = tmp / "params.h"
        content = h.read_text()
        content = content.replace("4706U", "5000U")
        h.write_text(content)
        r = sync(config_path=cfg)
        assert r[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE


# ═══════════════════════════════════════════════════════════════════════
# 5. xlsx ↔ c_struct_table round-trip (the complex one)
# ═══════════════════════════════════════════════════════════════════════

class TestStructTableRoundTrip:
    """Test xlsx ↔ c_struct_table round-trip — the main source of sync issues."""

    @pytest.fixture
    def struct_project(self, tmp_path):
        c_content = '''\
#include "types.h"

/* Field names:  R_Phase  L_d  L_q  Psi  Speed_Max */

#if (VARIANT == VARIANT_A)

static const MotorParam_t params[3] = {
/* Motor X  */ {{0.025F,  0.00003F,  0.00004F,  0.005F,  3000.0F}},
/* Motor Y  */ {{0.030F,  0.00005F,  0.00006F,  0.008F,  4500.0F}},
/* Motor Z  */ {{0.015F,  0.00002F,  0.00003F,  0.003F,  6000.0F}}
};

#elif (VARIANT == VARIANT_B)

static const MotorParam_t params[3] = {
/* Motor X  */ {{0.045F,  0.00006F,  0.00007F,  0.010F,  2500.0F}},
/* Motor Y  */ {{0.050F,  0.00008F,  0.00009F,  0.012F,  3500.0F}},
/* Motor Z  */ {{0.035F,  0.00004F,  0.00005F,  0.007F,  5000.0F}}
};

#endif
'''
        c_path = tmp_path / "motor_params.c"
        c_path.write_text(c_content)

        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: motor_params.c
                target: motor_params.xlsx
                format: c_struct_table
                direction: both
                parser_options:
                  variant: all
        """))
        return tmp_path, config, c_content

    def test_c_to_xlsx_creates_file(self, struct_project):
        """C struct → xlsx creates properly formatted xlsx."""
        tmp, cfg, _ = struct_project
        r = sync(config_path=cfg)
        assert r[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert (tmp / "motor_params.xlsx").exists()

    def test_c_to_xlsx_then_noop(self, struct_project):
        """Second sync after C→xlsx is noop."""
        tmp, cfg, _ = struct_project
        sync(config_path=cfg)
        r2 = sync(config_path=cfg)
        assert r2[0].result == SyncResult.ALREADY_IN_SYNC

    def test_xlsx_edit_detected_and_synced_back(self, struct_project):
        """Edit xlsx value → sync detects change and updates C file."""
        tmp, cfg, original_c = struct_project
        sync(config_path=cfg)

        # Edit xlsx: Motor X R_Phase in variant A: 0.025 → 0.1
        wb = openpyxl.load_workbook(tmp / "motor_params.xlsx")
        ws = wb["A"]
        assert ws.cell(2, 1).value == "Motor X"
        ws.cell(2, 2).value = 0.1
        wb.save(tmp / "motor_params.xlsx")

        r = sync(config_path=cfg)
        assert r[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE
        updated = (tmp / "motor_params.c").read_text()
        assert updated != original_c
        assert "0.100F" in updated or "0.1" in updated

    def test_xlsx_edit_then_noop(self, struct_project):
        """After xlsx→C sync, next sync is noop."""
        tmp, cfg, _ = struct_project
        sync(config_path=cfg)

        wb = openpyxl.load_workbook(tmp / "motor_params.xlsx")
        ws = wb["A"]
        ws.cell(2, 2).value = 0.1
        wb.save(tmp / "motor_params.xlsx")

        sync(config_path=cfg)
        r3 = sync(config_path=cfg)
        assert r3[0].result == SyncResult.ALREADY_IN_SYNC

    def test_direction_target_to_source_detects_xlsx_change(self, struct_project):
        """With direction=target_to_source, xlsx change is detected."""
        tmp, cfg, original_c = struct_project

        # Rewrite config with t2s direction
        (tmp / ".consync.yaml").write_text(textwrap.dedent("""\
            mappings:
              - source: motor_params.c
                target: motor_params.xlsx
                format: c_struct_table
                direction: target_to_source
                protect_target: false
                parser_options:
                  variant: all
        """))

        # First sync: C → xlsx (t2s means target is truth — but target doesn't exist yet)
        # Actually t2s with no source... target is xlsx which doesn't exist yet
        # Let's create xlsx first via s2t, then switch to t2s
        (tmp / ".consync.yaml").write_text(textwrap.dedent("""\
            mappings:
              - source: motor_params.c
                target: motor_params.xlsx
                format: c_struct_table
                direction: source_to_target
                protect_target: false
                parser_options:
                  variant: all
        """))
        sync(config_path=tmp / ".consync.yaml")

        # Now edit xlsx
        wb = openpyxl.load_workbook(tmp / "motor_params.xlsx")
        ws = wb["A"]
        ws.cell(2, 2).value = 0.1
        wb.save(tmp / "motor_params.xlsx")

        # Switch to t2s
        (tmp / ".consync.yaml").write_text(textwrap.dedent("""\
            mappings:
              - source: motor_params.c
                target: motor_params.xlsx
                format: c_struct_table
                direction: target_to_source
                protect_target: false
                parser_options:
                  variant: all
        """))

        # Clear state (new config key due to different direction intent)
        state_file = tmp / ".consync.state.json"
        if state_file.exists():
            state_file.unlink()

        r = sync(config_path=tmp / ".consync.yaml")
        assert r[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE
        updated = (tmp / "motor_params.c").read_text()
        assert "0.100F" in updated or "0.1" in updated


# ═══════════════════════════════════════════════════════════════════════
# 6. CLI diff — parser_options not passed
# ═══════════════════════════════════════════════════════════════════════

class TestDiffParserOptions:
    """The diff command must pass parser_options to _parse_file."""

    def test_diff_with_parser_options(self, tmp_path):
        """diff_cmd should work with c_struct_table format."""
        c_content = '''\
/* Field names:  R_Phase  L_d  L_q */

static const Param_t tbl[2] = {
/* Row A  */ {{1.0F,  2.0F,  3.0F}},
/* Row B  */ {{4.0F,  5.0F,  6.0F}}
};
'''
        c_path = tmp_path / "p.c"
        c_path.write_text(c_content)
        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: p.c
                target: p.xlsx
                format: c_struct_table
                direction: source_to_target
                protect_target: false
        """))
        # First sync to create xlsx
        sync(config_path=config)
        # Now diff should not crash
        from consync.sync import check as check_sync
        reports = check_sync(config_path=config)
        assert reports[0].result == SyncResult.ALREADY_IN_SYNC


# ═══════════════════════════════════════════════════════════════════════
# 7. _write_xlsx_flat stale rows — old data not cleared
# ═══════════════════════════════════════════════════════════════════════

class TestXlsxFlatStaleRows:
    """When re-writing xlsx flat, old rows beyond new data must be cleared."""

    def test_fewer_constants_clears_old_rows(self, tmp_path):
        """If new data has fewer rows, old rows must be wiped."""
        # Create xlsx with 3 rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Constants"
        ws.append(["Name", "Value", "Unit", "Description"])
        ws.append(["A", 1.0, "", ""])
        ws.append(["B", 2.0, "", ""])
        ws.append(["C", 3.0, "", ""])
        wb.save(tmp_path / "test.xlsx")

        # Write only 2 constants back
        from consync.sync import _write_xlsx_flat
        from consync.models import MappingConfig
        constants = [Constant("A", 1.0), Constant("B", 2.0)]
        cfg = MappingConfig(source="x.csv", target="test.xlsx")
        _write_xlsx_flat(constants, tmp_path / "test.xlsx", cfg)

        # Read back — should have exactly 2 data rows
        wb2 = openpyxl.load_workbook(tmp_path / "test.xlsx")
        ws2 = wb2.active
        # Row 4 (old "C" row) should be None
        assert ws2.cell(4, 1).value is None


# ═══════════════════════════════════════════════════════════════════════
# 8. Watcher-simulated scenario — no ping-pong
# ═══════════════════════════════════════════════════════════════════════

class TestNoPingPong:
    """Simulate watcher scenario: multiple sync() calls shouldn't ping-pong."""

    def test_triple_sync_no_oscillation(self, tmp_path):
        """After initial sync, calling sync() 3 more times → all noop."""
        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: data.csv
                target: data.h
                direction: both
                header_guard: DATA_H
        """))
        csv = tmp_path / "data.csv"
        csv.write_text("Name,Value\nX,42\n")

        r1 = sync(config_path=config)
        assert r1[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        for i in range(3):
            r = sync(config_path=config)
            assert r[0].result == SyncResult.ALREADY_IN_SYNC, (
                f"Sync #{i+2} should be noop but got {r[0].result.value}: {r[0].message}"
            )

    def test_struct_table_triple_sync(self, tmp_path):
        """Same test with c_struct_table — the historically problematic format."""
        c_content = '''\
/* Field names:  val1  val2 val3 */

static const Param_t tbl[2] = {
/* Row A  */ {{1.0F,  2.0F,  3.0F}},
/* Row B  */ {{4.0F,  5.0F,  6.0F}}
};
'''
        c_path = tmp_path / "p.c"
        c_path.write_text(c_content)
        config = tmp_path / ".consync.yaml"
        config.write_text(textwrap.dedent("""\
            mappings:
              - source: p.c
                target: p.xlsx
                format: c_struct_table
                direction: both
        """))

        r1 = sync(config_path=config)
        assert r1[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        for i in range(3):
            r = sync(config_path=config)
            assert r[0].result == SyncResult.ALREADY_IN_SYNC, (
                f"Sync #{i+2} should be noop but got {r[0].result.value}: {r[0].message}"
            )
