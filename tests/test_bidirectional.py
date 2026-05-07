"""Tests for bidirectional sync: xlsx table parser + c_struct_table renderer round-trip."""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest

from consync.models import Constant, MappingConfig, SyncDirection


@pytest.fixture
def tmp_dir(tmp_path):
    """Create a temporary directory for test files."""
    return tmp_path


@pytest.fixture
def simple_c_file(tmp_dir):
    """Create a simple C struct table file for testing."""
    content = '''\
#include "types.h"

/* Field names:  R_Phase  L_d  L_q  Psi  Speed_Max */

#if (VARIANT == VARIANT_A)

static const MotorParam_t params[3] = {
/* Motor X  */ {{0.025F,  0.00003F,  0.00004F,  0.005F,  3000.0F}},
/* Motor Y  */ {{0.030F,  0.00005F,  0.00006F,  0.008F,  4500.0F}},
/* Motor Z  */ {{0.015F,  0.00002F,  0.00003F,  0.003F,  6000.0F}}
};

#elif (VARIANT == VARIANT_B)

static const MotorParam_t params[3] = {
/* Motor X  */ {{0.045F,  0.00006F,  0.00007F,  0.010F,  2500.0F}},
/* Motor Y  */ {{0.050F,  0.00008F,  0.00009F,  0.012F,  3500.0F}},
/* Motor Z  */ {{0.035F,  0.00004F,  0.00005F,  0.007F,  5000.0F}}
};

#endif
'''
    filepath = tmp_dir / "motor_params.c"
    filepath.write_text(content)
    return filepath


@pytest.fixture
def config(simple_c_file, tmp_dir):
    """Create a MappingConfig for testing."""
    return MappingConfig(
        source=str(simple_c_file),
        target=str(tmp_dir / "motor_params.xlsx"),
        source_format="c_struct_table",
        target_format="xlsx",
        direction=SyncDirection.BOTH,
        parser_options={"variant": "all"},
    )


class TestXlsxTableParser:
    """Test xlsx parser with table-layout format."""

    def test_parse_table_layout(self, simple_c_file, tmp_dir, config):
        """Parse C → write xlsx → parse xlsx and compare."""
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        # Parse C file
        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        assert len(c_constants) == 30  # 3 motors × 5 fields × 2 variants

        # Write xlsx
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)
        assert xlsx_path.exists()

        # Parse xlsx back
        xl_constants = parse_xlsx(xlsx_path)
        assert len(xl_constants) == 30

        # Verify metadata is reconstructed
        first = xl_constants[0]
        assert first.metadata["row_label"] is not None
        assert first.metadata["field"] is not None
        assert first.metadata["field_index"] is not None
        assert first.metadata["variant"] is not None

    def test_values_preserved_roundtrip(self, simple_c_file, tmp_dir, config):
        """Values survive C → xlsx → parse cycle."""
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)

        xl_constants = parse_xlsx(xlsx_path)

        # Build value maps
        c_map = {(c.metadata["row_label"], c.metadata["field_index"]): c.value for c in c_constants}
        xl_map = {(c.metadata["row_label"], c.metadata["field_index"]): c.value for c in xl_constants}

        for key in c_map:
            c_val = c_map[key]
            xl_val = xl_map.get(key)
            if isinstance(c_val, float) and isinstance(xl_val, float):
                assert abs(c_val - xl_val) < abs(c_val) * 1e-6 + 1e-15, (
                    f"Mismatch at {key}: C={c_val}, XL={xl_val}"
                )
            else:
                assert c_val == xl_val, f"Mismatch at {key}: C={c_val}, XL={xl_val}"

    def test_auto_detects_table_layout(self, simple_c_file, tmp_dir, config):
        """Parser auto-detects table vs flat layout."""
        import openpyxl
        from consync.parsers.xlsx import parse_xlsx, _is_table_layout
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.sync import _write_xlsx

        # Write table layout
        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)

        # Check detection
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]
        assert _is_table_layout(ws)

    def test_flat_layout_still_works(self, tmp_dir):
        """Flat xlsx format still parsed correctly."""
        import openpyxl
        from consync.parsers.xlsx import parse_xlsx

        # Create flat xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value", "Unit", "Description"])
        ws.append(["PI", 3.14159, "rad", "Pi constant"])
        ws.append(["SPEED_MAX", 5000, "rpm", "Max speed"])
        xlsx_path = tmp_dir / "flat.xlsx"
        wb.save(xlsx_path)

        constants = parse_xlsx(str(xlsx_path))
        assert len(constants) == 2
        assert constants[0].name == "PI"
        assert abs(constants[0].value - 3.14159) < 1e-10
        assert constants[1].name == "SPEED_MAX"
        assert constants[1].value == 5000

    def test_multi_sheet_variants(self, simple_c_file, tmp_dir, config):
        """Multiple sheets map to multiple variants."""
        import openpyxl
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)

        xl_constants = parse_xlsx(xlsx_path)

        # Check that both variants are present
        variants = set(c.metadata["variant"] for c in xl_constants)
        assert "A" in variants
        assert "B" in variants


class TestBidirectionalSync:
    """Test complete bidirectional sync from xlsx → C file."""

    def test_single_value_change(self, simple_c_file, tmp_dir, config):
        """Changing one value in xlsx updates only that value in C."""
        import openpyxl
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.renderers.c_struct_table import render_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        # Step 1: C → xlsx
        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)

        # Keep original content for comparison
        original = simple_c_file.read_text()

        # Step 2: Edit one value in xlsx (Motor X / R_Phase in VARIANT_A: 0.025 → 0.1)
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["A"]
        # Row 2 = Motor X, Col 2 = R_Phase
        assert ws.cell(2, 1).value == "Motor X"
        assert abs(ws.cell(2, 2).value - 0.025) < 1e-10
        ws.cell(2, 2).value = 0.1
        wb.save(xlsx_path)

        # Step 3: xlsx → C (render)
        xl_constants = parse_xlsx(xlsx_path)
        render_c_struct_table(xl_constants, simple_c_file, config=config)

        # Step 4: Verify
        updated = simple_c_file.read_text()
        assert updated != original  # Something changed
        assert "0.100F" in updated or "1.00E-01F" in updated  # New value present
        # Check original value is gone from that line
        lines_orig = original.splitlines()
        lines_new = updated.splitlines()

        # Count differences
        diffs = [(i, a, b) for i, (a, b) in enumerate(zip(lines_orig, lines_new)) if a != b]
        assert len(diffs) == 1, f"Expected 1 line changed, got {len(diffs)}"

        # The changed line should be "Motor X" in VARIANT_A section
        changed_line = diffs[0][2]
        assert "Motor X" in changed_line

    def test_multiple_value_changes(self, simple_c_file, tmp_dir, config):
        """Changing multiple values in xlsx updates all correctly."""
        import openpyxl
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.renderers.c_struct_table import render_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        # Step 1: C → xlsx
        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)
        original = simple_c_file.read_text()

        # Step 2: Edit multiple values
        wb = openpyxl.load_workbook(xlsx_path)
        # Change Motor Y / R_Phase in VARIANT_A
        ws_a = wb["A"]
        ws_a.cell(3, 2).value = 0.099  # Motor Y R_Phase
        # Change Motor X / Speed_Max in VARIANT_B
        ws_b = wb["B"]
        ws_b.cell(2, 6).value = 9999.0  # Motor X Speed_Max
        wb.save(xlsx_path)

        # Step 3: Sync back
        xl_constants = parse_xlsx(xlsx_path)
        render_c_struct_table(xl_constants, simple_c_file, config=config)

        # Step 4: Verify
        updated = simple_c_file.read_text()
        lines_orig = original.splitlines()
        lines_new = updated.splitlines()
        diffs = [(i, a, b) for i, (a, b) in enumerate(zip(lines_orig, lines_new)) if a != b]
        assert len(diffs) == 2, f"Expected 2 lines changed, got {len(diffs)}"

    def test_no_change_no_write(self, simple_c_file, tmp_dir, config):
        """If xlsx matches C file, no changes are made."""
        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.renderers.c_struct_table import render_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        # C → xlsx
        c_constants = parse_c_struct_table(simple_c_file, variant="all")
        xlsx_path = tmp_dir / "motor_params.xlsx"
        _write_xlsx(c_constants, xlsx_path, config)

        original = simple_c_file.read_text()

        # xlsx → C (no changes)
        xl_constants = parse_xlsx(xlsx_path)
        render_c_struct_table(xl_constants, simple_c_file, config=config)

        # File unchanged
        assert simple_c_file.read_text() == original

    def test_expression_values_preserved(self, tmp_dir):
        """Expression/macro values in C are not modified."""
        content = '''\
/* Fields: val1 val2 val3 */

static const Param_t tbl[2] = {
/* Row A  */ {{1.0F,  MY_MACRO,  3.0F}},
/* Row B  */ {{4.0F,  OTHER_DEF, 6.0F}}
};
'''
        c_path = tmp_dir / "expr_test.c"
        c_path.write_text(content)

        cfg = MappingConfig(
            source=str(c_path),
            target=str(tmp_dir / "expr.xlsx"),
            source_format="c_struct_table",
            target_format="xlsx",
            direction=SyncDirection.BOTH,
            parser_options={},
        )

        from consync.parsers.c_struct_table import parse_c_struct_table
        from consync.renderers.c_struct_table import render_c_struct_table
        from consync.parsers.xlsx import parse_xlsx
        from consync.sync import _write_xlsx

        # C → xlsx → C round-trip
        c_constants = parse_c_struct_table(c_path)
        xlsx_path = tmp_dir / "expr.xlsx"
        _write_xlsx(c_constants, xlsx_path, cfg)

        xl_constants = parse_xlsx(xlsx_path)
        render_c_struct_table(xl_constants, c_path, config=cfg)

        # Expressions should be untouched
        result = c_path.read_text()
        assert "MY_MACRO" in result
        assert "OTHER_DEF" in result


class TestXlsxTableDetection:
    """Test table layout detection edge cases."""

    def test_not_table_when_name_value_headers(self, tmp_dir):
        """Should NOT detect table layout when headers are Name/Value."""
        import openpyxl
        from consync.parsers.xlsx import _is_table_layout

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Name"
        ws.cell(1, 2).value = "Value"
        assert not _is_table_layout(ws)

    def test_detects_motor_variant_header(self, tmp_dir):
        """Detects table layout from 'Motor Variant' header."""
        import openpyxl
        from consync.parsers.xlsx import _is_table_layout

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Motor Variant"
        ws.cell(1, 2).value = "R_Phase"
        assert _is_table_layout(ws)

    def test_detects_variant_header(self, tmp_dir):
        """Detects table layout from 'Variant' header."""
        import openpyxl
        from consync.parsers.xlsx import _is_table_layout

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Variant"
        ws.cell(1, 2).value = "Param1"
        assert _is_table_layout(ws)

    def test_empty_header_not_table(self, tmp_dir):
        """Empty first header → not table layout."""
        import openpyxl
        from consync.parsers.xlsx import _is_table_layout

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = None
        assert not _is_table_layout(ws)
