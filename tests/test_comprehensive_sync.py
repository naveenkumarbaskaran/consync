"""Comprehensive integration tests for all consync sync scenarios.

Covers:
- Source file creation (target doesn't exist yet)
- One-way sync (source_to_target / target_to_source)
- Bidirectional sync (both directions, conflict detection)
- All supported format combinations
- Edge cases (empty files, special values, precision)
"""

import textwrap
from pathlib import Path

import pytest

from consync.models import Constant, MappingConfig, SyncDirection
from consync.sync import sync, check, SyncResult


# ============================================================================
# Fixtures — Source file content for each format
# ============================================================================


C_HEADER_CONTENT = textwrap.dedent("""\
    #ifndef PARAMS_H
    #define PARAMS_H

    static const double SPEED_MAX = 3000.0;  /* rpm — Maximum motor speed */
    static const double R_PHASE = 0.025;  /* Ohm — Phase resistance */
    static const uint32_t POLE_PAIRS = 5u;  /* — Number of pole pairs */
    #define VOLTAGE_LIMIT (48.0)  /* V — Supply voltage limit */

    #endif
""")


CSV_CONTENT = textwrap.dedent("""\
    Name,Value,Unit,Description
    SPEED_MAX,3000.0,rpm,Maximum motor speed
    R_PHASE,0.025,Ohm,Phase resistance
    POLE_PAIRS,5,,Number of pole pairs
    VOLTAGE_LIMIT,48.0,V,Supply voltage limit
""")


JSON_CONTENT_FLAT = '{"SPEED_MAX": 3000.0, "R_PHASE": 0.025, "POLE_PAIRS": 5, "VOLTAGE_LIMIT": 48.0}'

JSON_CONTENT_STRUCTURED = textwrap.dedent("""\
    [
      {"name": "SPEED_MAX", "value": 3000.0, "unit": "rpm", "description": "Maximum motor speed"},
      {"name": "R_PHASE", "value": 0.025, "unit": "Ohm", "description": "Phase resistance"},
      {"name": "POLE_PAIRS", "value": 5, "unit": "", "description": "Number of pole pairs"},
      {"name": "VOLTAGE_LIMIT", "value": 48.0, "unit": "V", "description": "Supply voltage limit"}
    ]
""")


TOML_CONTENT = textwrap.dedent("""\
    [constants]
    SPEED_MAX = 3000.0
    R_PHASE = 0.025
    POLE_PAIRS = 5
    VOLTAGE_LIMIT = 48.0
""")

TOML_CONTENT_RICH = textwrap.dedent("""\
    [constants.SPEED_MAX]
    value = 3000.0
    unit = "rpm"
    description = "Maximum motor speed"

    [constants.R_PHASE]
    value = 0.025
    unit = "Ohm"
    description = "Phase resistance"

    [constants.POLE_PAIRS]
    value = 5
    unit = ""
    description = "Number of pole pairs"

    [constants.VOLTAGE_LIMIT]
    value = 48.0
    unit = "V"
    description = "Supply voltage limit"
""")


C_STRUCT_TABLE_CONTENT = textwrap.dedent("""\
    #include "types.h"

    /*  R_Phase  L_d  L_q  Psi  Speed_Max */

    static const MotorParam_t params[3] = {
    /* Motor X  */ {{0.025F,  0.00003F,  0.00004F,  0.005F,  3000.0F}},
    /* Motor Y  */ {{0.030F,  0.00005F,  0.00006F,  0.008F,  4500.0F}},
    /* Motor Z  */ {{0.015F,  0.00002F,  0.00003F,  0.003F,  6000.0F}}
    };
""")


C_STRUCT_TABLE_VARIANTS = textwrap.dedent("""\
    #include "types.h"

    /*  R_Phase  L_d  L_q  Psi  Speed_Max */

    #if (MOTOR == MOTOR_A)

    static const MotorParam_t params[2] = {
    /* Alpha  */ {{0.025F,  3.0E-5F,  4.0E-5F,  0.005F,  3000.0F}},
    /* Beta   */ {{0.030F,  5.0E-5F,  6.0E-5F,  0.008F,  4500.0F}}
    };

    #elif (MOTOR == MOTOR_B)

    static const MotorParam_t params[2] = {
    /* Alpha  */ {{0.045F,  6.0E-5F,  7.0E-5F,  0.010F,  2500.0F}},
    /* Beta   */ {{0.050F,  8.0E-5F,  9.0E-5F,  0.012F,  3500.0F}}
    };

    #endif
""")


# ============================================================================
# Helper functions
# ============================================================================


def create_config(tmp_path, source_name, target_name, direction="source_to_target",
                  source_format="", target_format="", parser_options=None):
    """Create a .consync.yaml config file."""
    cfg = f"""\
mappings:
  - source: {source_name}
    target: {target_name}
    direction: {direction}
"""
    if source_format:
        cfg += f"    source_format: {source_format}\n"
    if target_format:
        cfg += f"    target_format: {target_format}\n"
    if parser_options:
        cfg += "    parser_options:\n"
        for k, v in parser_options.items():
            cfg += f"      {k}: {v}\n"

    config_path = tmp_path / ".consync.yaml"
    config_path.write_text(cfg)
    return config_path


# ============================================================================
# Test Class: Source File Creation (target doesn't exist)
# ============================================================================


class TestSourceFileCreation:
    """Test that sync creates the target file from scratch for all format pairs."""

    def test_csv_to_c_header(self, tmp_path):
        """CSV source → C header target (first sync creates .h)."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 4

        target = tmp_path / "params.h"
        assert target.exists()
        content = target.read_text()
        assert "SPEED_MAX" in content
        assert "R_PHASE" in content
        assert "POLE_PAIRS" in content
        assert "VOLTAGE_LIMIT" in content

    def test_csv_to_json(self, tmp_path):
        """CSV source → JSON target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.json"
        assert target.exists()
        import json
        data = json.loads(target.read_text())
        assert "constants" in data
        assert len(data["constants"]) == 4

    def test_csv_to_python(self, tmp_path):
        """CSV source → Python module target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.py")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.py"
        assert target.exists()
        content = target.read_text()
        assert "SPEED_MAX" in content
        assert "3000" in content

    def test_csv_to_rust(self, tmp_path):
        """CSV source → Rust const target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.rs")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.rs"
        assert target.exists()
        content = target.read_text()
        assert "pub const SPEED_MAX" in content
        assert "f64" in content or "i64" in content

    def test_csv_to_verilog(self, tmp_path):
        """CSV source → Verilog parameter target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.v")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.v"
        assert target.exists()
        content = target.read_text()
        assert "parameter" in content
        assert "SPEED_MAX" in content

    def test_csv_to_vhdl(self, tmp_path):
        """CSV source → VHDL constant target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.vhd")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.vhd"
        assert target.exists()
        content = target.read_text()
        assert "constant" in content.lower()
        assert "SPEED_MAX" in content

    def test_csv_to_csharp(self, tmp_path):
        """CSV source → C# const target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.cs")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.cs"
        assert target.exists()
        content = target.read_text()
        assert "const" in content
        assert "SPEED_MAX" in content

    def test_csv_to_csv_roundtrip(self, tmp_path):
        """CSV source → CSV target (copy/reformat)."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "output.csv",
                               target_format="csv")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "output.csv"
        assert target.exists()
        content = target.read_text()
        assert "SPEED_MAX" in content

    def test_csv_to_xlsx(self, tmp_path):
        """CSV source → Excel target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.xlsx")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        target = tmp_path / "params.xlsx"
        assert target.exists()

    def test_json_to_c_header(self, tmp_path):
        """JSON source → C header target."""
        (tmp_path / "params.json").write_text(JSON_CONTENT_STRUCTURED)
        config = create_config(tmp_path, "params.json", "params.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 4

        content = (tmp_path / "params.h").read_text()
        assert "SPEED_MAX" in content

    def test_toml_to_c_header(self, tmp_path):
        """TOML source → C header target."""
        (tmp_path / "params.toml").write_text(TOML_CONTENT)
        config = create_config(tmp_path, "params.toml", "params.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 4

    def test_c_header_to_csv(self, tmp_path):
        """C header source → CSV target."""
        (tmp_path / "params.h").write_text(C_HEADER_CONTENT)
        config = create_config(tmp_path, "params.h", "params.csv")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        content = (tmp_path / "params.csv").read_text()
        assert "SPEED_MAX" in content
        assert "3000" in content

    def test_c_struct_table_to_xlsx(self, tmp_path):
        """C struct table → Excel (table layout with variants)."""
        (tmp_path / "motor.c").write_text(C_STRUCT_TABLE_VARIANTS)
        config = create_config(
            tmp_path, "motor.c", "motor.xlsx",
            source_format="c_struct_table",
            parser_options={"variant": "all"},
        )

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count > 0

        assert (tmp_path / "motor.xlsx").exists()

    def test_second_sync_is_noop(self, tmp_path):
        """After initial sync, second sync detects already in sync."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h")

        sync(config_path=config)
        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.ALREADY_IN_SYNC


# ============================================================================
# Test Class: One-Way Sync (source_to_target only)
# ============================================================================


class TestOneWaySourceToTarget:
    """Test one-way sync where source is always the authoritative copy."""

    def test_source_change_syncs_to_target(self, tmp_path):
        """Modifying source causes re-sync to target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h",
                               direction="source_to_target")

        # Initial sync
        sync(config_path=config)
        original_h = (tmp_path / "params.h").read_text()

        # Modify source
        new_csv = CSV_CONTENT.replace("3000.0", "5000.0")
        (tmp_path / "params.csv").write_text(new_csv)

        # Re-sync
        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        updated_h = (tmp_path / "params.h").read_text()
        assert "5000" in updated_h
        assert updated_h != original_h

    def test_target_change_ignored_in_source_to_target(self, tmp_path):
        """In source_to_target mode, target changes are overwritten on next sync."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="source_to_target")

        # Initial sync
        sync(config_path=config)

        # Modify target (should be ignored/overwritten)
        (tmp_path / "params.json").write_text('{"constants": []}')

        # Modify source to trigger re-sync
        new_csv = CSV_CONTENT.replace("3000.0", "7777.0")
        (tmp_path / "params.csv").write_text(new_csv)

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        import json
        data = json.loads((tmp_path / "params.json").read_text())
        assert any(c["value"] == 7777.0 for c in data["constants"])

    def test_csv_to_multiple_targets(self, tmp_path):
        """One source can sync to multiple targets in separate mappings."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.h
                direction: source_to_target
              - source: params.csv
                target: params.json
                direction: source_to_target
              - source: params.csv
                target: params.py
                direction: source_to_target
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        reports = sync(config_path=config)
        assert len(reports) == 3
        assert all(r.result == SyncResult.SYNCED_SOURCE_TO_TARGET for r in reports)
        assert (tmp_path / "params.h").exists()
        assert (tmp_path / "params.json").exists()
        assert (tmp_path / "params.py").exists()

    def test_json_flat_to_csv(self, tmp_path):
        """JSON flat format → CSV."""
        (tmp_path / "params.json").write_text(JSON_CONTENT_FLAT)
        config = create_config(tmp_path, "params.json", "params.csv",
                               direction="source_to_target")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp_path / "params.csv").read_text()
        assert "SPEED_MAX" in content

    def test_json_structured_to_c_header(self, tmp_path):
        """JSON structured format → C header."""
        (tmp_path / "params.json").write_text(JSON_CONTENT_STRUCTURED)
        config = create_config(tmp_path, "params.json", "params.h",
                               direction="source_to_target")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp_path / "params.h").read_text()
        assert "SPEED_MAX" in content
        assert "R_PHASE" in content

    def test_toml_rich_to_c_header(self, tmp_path):
        """TOML with unit/description → C header."""
        (tmp_path / "params.toml").write_text(TOML_CONTENT_RICH)
        config = create_config(tmp_path, "params.toml", "params.h",
                               direction="source_to_target")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp_path / "params.h").read_text()
        assert "SPEED_MAX" in content
        # Description should appear as comment
        assert "motor speed" in content.lower() or "Maximum" in content

    def test_c_header_to_rust(self, tmp_path):
        """C header → Rust constants."""
        (tmp_path / "params.h").write_text(C_HEADER_CONTENT)
        config = create_config(tmp_path, "params.h", "params.rs",
                               direction="source_to_target")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp_path / "params.rs").read_text()
        assert "pub const" in content
        assert "SPEED_MAX" in content

    def test_c_header_to_python(self, tmp_path):
        """C header → Python module."""
        (tmp_path / "params.h").write_text(C_HEADER_CONTENT)
        config = create_config(tmp_path, "params.h", "params.py",
                               direction="source_to_target")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        content = (tmp_path / "params.py").read_text()
        assert "SPEED_MAX" in content
        assert "3000" in content


class TestOneWayTargetToSource:
    """Test one-way sync where target is the authoritative copy."""

    def test_target_to_source_creates_source(self, tmp_path):
        """In t2s mode, target creates source on first sync."""
        # Target exists, source doesn't
        (tmp_path / "params.json").write_text(JSON_CONTENT_STRUCTURED)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="target_to_source")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE
        assert (tmp_path / "params.csv").exists()
        content = (tmp_path / "params.csv").read_text()
        assert "SPEED_MAX" in content

    def test_target_change_syncs_to_source(self, tmp_path):
        """Modifying target updates source in t2s mode."""
        (tmp_path / "params.json").write_text(JSON_CONTENT_STRUCTURED)
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="target_to_source")

        # Initial sync
        sync(config_path=config)

        # Modify target
        new_json = JSON_CONTENT_STRUCTURED.replace("3000.0", "9999.0")
        (tmp_path / "params.json").write_text(new_json)

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE
        content = (tmp_path / "params.csv").read_text()
        assert "9999" in content


# ============================================================================
# Test Class: Bidirectional Sync
# ============================================================================


class TestBidirectionalSync:
    """Test bidirectional sync with conflict detection."""

    def test_source_change_detected(self, tmp_path):
        """In both mode, source change triggers source → target sync."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="both")

        # Initial sync
        sync(config_path=config)

        # Modify source only
        new_csv = CSV_CONTENT.replace("3000.0", "4000.0")
        (tmp_path / "params.csv").write_text(new_csv)

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        import json
        data = json.loads((tmp_path / "params.json").read_text())
        values = [c["value"] for c in data["constants"]]
        assert 4000.0 in values

    def test_target_change_detected(self, tmp_path):
        """In both mode, target change triggers target → source sync."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="both")

        # Initial sync
        sync(config_path=config)

        # Modify target only
        import json
        data = json.loads((tmp_path / "params.json").read_text())
        data["constants"][0]["value"] = 8888.0
        (tmp_path / "params.json").write_text(json.dumps(data, indent=2))

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE

        content = (tmp_path / "params.csv").read_text()
        assert "8888" in content

    def test_both_changed_conflict(self, tmp_path):
        """Both files changed → conflict detected."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.json
                direction: both
            on_conflict: fail
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        # Initial sync
        sync(config_path=config)

        # Modify BOTH files
        new_csv = CSV_CONTENT.replace("3000.0", "1111.0")
        (tmp_path / "params.csv").write_text(new_csv)

        import json
        data = json.loads((tmp_path / "params.json").read_text())
        data["constants"][0]["value"] = 2222.0
        (tmp_path / "params.json").write_text(json.dumps(data, indent=2))

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.CONFLICT

    def test_conflict_source_wins(self, tmp_path):
        """on_conflict: source_wins resolves conflict by using source."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.json
                direction: both
            on_conflict: source_wins
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)

        # Modify both
        new_csv = CSV_CONTENT.replace("3000.0", "1111.0")
        (tmp_path / "params.csv").write_text(new_csv)

        import json
        data = json.loads((tmp_path / "params.json").read_text())
        data["constants"][0]["value"] = 2222.0
        (tmp_path / "params.json").write_text(json.dumps(data, indent=2))

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        data = json.loads((tmp_path / "params.json").read_text())
        values = [c["value"] for c in data["constants"]]
        assert 1111.0 in values  # Source wins

    def test_conflict_target_wins(self, tmp_path):
        """on_conflict: target_wins resolves conflict by using target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.json
                direction: both
            on_conflict: target_wins
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)

        # Modify both
        new_csv = CSV_CONTENT.replace("3000.0", "1111.0")
        (tmp_path / "params.csv").write_text(new_csv)

        import json
        data = json.loads((tmp_path / "params.json").read_text())
        data["constants"][0]["value"] = 2222.0
        (tmp_path / "params.json").write_text(json.dumps(data, indent=2))

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE

        content = (tmp_path / "params.csv").read_text()
        assert "2222" in content  # Target wins


# ============================================================================
# Test Class: c_struct_table Bidirectional (xlsx ↔ C)
# ============================================================================


class TestCStructTableBidirectional:
    """Test bidirectional sync between C struct table files and Excel."""

    def test_c_to_xlsx_creates_table_layout(self, tmp_path):
        """C struct table → Excel creates multi-sheet table layout."""
        (tmp_path / "motor.c").write_text(C_STRUCT_TABLE_VARIANTS)
        config = create_config(
            tmp_path, "motor.c", "motor.xlsx",
            source_format="c_struct_table",
            direction="both",
            parser_options={"variant": "all"},
        )

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

        import openpyxl
        wb = openpyxl.load_workbook(tmp_path / "motor.xlsx")
        # Should have variant sheets + Info
        sheet_names = [s for s in wb.sheetnames if s != "Info"]
        assert len(sheet_names) == 2  # Two variants

    def test_xlsx_edit_syncs_back_to_c(self, tmp_path):
        """Editing value in Excel syncs back to C file correctly."""
        (tmp_path / "motor.c").write_text(C_STRUCT_TABLE_VARIANTS)
        config = create_config(
            tmp_path, "motor.c", "motor.xlsx",
            source_format="c_struct_table",
            direction="both",
            parser_options={"variant": "all"},
        )

        # Initial sync C → xlsx
        sync(config_path=config)

        original_c = (tmp_path / "motor.c").read_text()

        # Edit xlsx: change Alpha R_Phase in first variant sheet
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path / "motor.xlsx")
        # Get first data sheet (not Info)
        data_sheets = [s for s in wb.sheetnames if s != "Info"]
        ws = wb[data_sheets[0]]

        # Row 2 = first motor, col 2 = first field (R_Phase)
        row_label = ws.cell(2, 1).value
        assert row_label is not None
        original_val = ws.cell(2, 2).value
        ws.cell(2, 2).value = 0.999  # Change first field value
        wb.save(tmp_path / "motor.xlsx")

        # Sync xlsx → C
        reports = sync(config_path=config, force_direction="target")
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE

        # Verify C file changed
        updated_c = (tmp_path / "motor.c").read_text()
        assert updated_c != original_c

        # Verify specific change: the row should have new value
        found = False
        for line in updated_c.splitlines():
            if row_label in line and "999" in line:
                found = True
                break
        assert found, f"Expected 0.999 in line with '{row_label}'"

    def test_xlsx_no_change_preserves_c(self, tmp_path):
        """If xlsx not modified, C file stays identical."""
        (tmp_path / "motor.c").write_text(C_STRUCT_TABLE_VARIANTS)
        config = create_config(
            tmp_path, "motor.c", "motor.xlsx",
            source_format="c_struct_table",
            direction="both",
            parser_options={"variant": "all"},
        )

        sync(config_path=config)
        original_c = (tmp_path / "motor.c").read_text()

        # Force re-sync from xlsx (no changes)
        reports = sync(config_path=config, force_direction="target")
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE

        # C file should be identical
        assert (tmp_path / "motor.c").read_text() == original_c

    def test_single_variant_roundtrip(self, tmp_path):
        """Single variant (no #if blocks) round-trips cleanly."""
        (tmp_path / "motor.c").write_text(C_STRUCT_TABLE_CONTENT)
        config = create_config(
            tmp_path, "motor.c", "motor.xlsx",
            source_format="c_struct_table",
            direction="both",
        )

        # C → xlsx
        sync(config_path=config)
        assert (tmp_path / "motor.xlsx").exists()

        original_c = (tmp_path / "motor.c").read_text()

        # xlsx → C (no changes)
        reports = sync(config_path=config, force_direction="target")
        assert (tmp_path / "motor.c").read_text() == original_c


# ============================================================================
# Test Class: Cross-Format Value Preservation
# ============================================================================


class TestValuePreservation:
    """Test that values are preserved accurately across format conversions."""

    def test_integers_preserved(self, tmp_path):
        """Integer values stay as integers through conversion."""
        csv_content = "Name,Value,Unit,Description\nCOUNT,42,,An integer\nMAX,255,,Byte max\n"
        (tmp_path / "data.csv").write_text(csv_content)
        config = create_config(tmp_path, "data.csv", "data.json",
                               direction="both")

        sync(config_path=config)

        import json
        data = json.loads((tmp_path / "data.json").read_text())
        values = {c["name"]: c["value"] for c in data["constants"]}
        assert values["COUNT"] == 42
        assert values["MAX"] == 255
        assert isinstance(values["COUNT"], int)

    def test_floats_preserved(self, tmp_path):
        """Float precision is maintained through conversion."""
        csv_content = "Name,Value,Unit,Description\nPI,3.14159265358979,,Pi\nE,2.71828182845905,,Euler\n"
        (tmp_path / "data.csv").write_text(csv_content)
        config = create_config(tmp_path, "data.csv", "data.json",
                               direction="both")

        sync(config_path=config)

        import json
        data = json.loads((tmp_path / "data.json").read_text())
        values = {c["name"]: c["value"] for c in data["constants"]}
        assert abs(values["PI"] - 3.14159265358979) < 1e-12
        assert abs(values["E"] - 2.71828182845905) < 1e-12

    def test_scientific_notation_preserved_in_c(self, tmp_path):
        """Scientific notation values work in C struct tables."""
        c_content = textwrap.dedent("""\
            /*  Resistance  Inductance */

            static const Param_t tbl[2] = {
            /* Sensor A  */ {{1.5E-3F,  2.0E-6F}},
            /* Sensor B  */ {{3.0E-3F,  4.5E-6F}}
            };
        """)
        (tmp_path / "sci.c").write_text(c_content)
        config = create_config(
            tmp_path, "sci.c", "sci.xlsx",
            source_format="c_struct_table",
            direction="both",
        )

        sync(config_path=config)
        original = (tmp_path / "sci.c").read_text()

        # Round-trip should preserve
        sync(config_path=config, force_direction="target")
        assert (tmp_path / "sci.c").read_text() == original

    def test_zero_values_preserved(self, tmp_path):
        """Zero values don't get dropped or mangled."""
        csv_content = "Name,Value,Unit,Description\nOFFSET,0.0,,Zero offset\nBIAS,0,,Zero bias\n"
        (tmp_path / "data.csv").write_text(csv_content)
        config = create_config(tmp_path, "data.csv", "data.json",
                               direction="both")

        sync(config_path=config)

        import json
        data = json.loads((tmp_path / "data.json").read_text())
        values = {c["name"]: c["value"] for c in data["constants"]}
        assert values["OFFSET"] == 0.0
        assert values["BIAS"] == 0

    def test_negative_values_preserved(self, tmp_path):
        """Negative values preserved through conversion."""
        csv_content = "Name,Value,Unit,Description\nTEMP_MIN,-40.0,C,Min temp\nOFFSET,-0.005,V,Neg offset\n"
        (tmp_path / "data.csv").write_text(csv_content)
        config = create_config(tmp_path, "data.csv", "data.h",
                               direction="source_to_target")

        sync(config_path=config)
        content = (tmp_path / "data.h").read_text()
        assert "-40" in content
        assert "-0.005" in content


# ============================================================================
# Test Class: Force Direction Override
# ============================================================================


class TestForceDirection:
    """Test the force_direction parameter."""

    def test_force_source_overrides_config(self, tmp_path):
        """force_direction='source' overrides any config direction."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h",
                               direction="target_to_source")

        # Even though direction is t2s, force source
        reports = sync(config_path=config, force_direction="source")
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert (tmp_path / "params.h").exists()

    def test_force_target_overrides_config(self, tmp_path):
        """force_direction='target' overrides any config direction."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        (tmp_path / "params.h").write_text(C_HEADER_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h",
                               direction="source_to_target")

        # Force target even though config says s2t
        reports = sync(config_path=config, force_direction="target")
        assert reports[0].result == SyncResult.SYNCED_TARGET_TO_SOURCE


# ============================================================================
# Test Class: Dry Run
# ============================================================================


class TestDryRun:
    """Test dry run mode doesn't write files."""

    def test_dry_run_no_file_created(self, tmp_path):
        """Dry run reports what would happen without creating files."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h")

        reports = sync(config_path=config, dry_run=True)
        assert "DRY RUN" in reports[0].message
        assert not (tmp_path / "params.h").exists()

    def test_dry_run_no_file_modified(self, tmp_path):
        """Dry run doesn't modify existing target."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h")

        # First real sync
        sync(config_path=config)
        original = (tmp_path / "params.h").read_text()

        # Modify source
        new_csv = CSV_CONTENT.replace("3000.0", "9999.0")
        (tmp_path / "params.csv").write_text(new_csv)

        # Dry run
        reports = sync(config_path=config, dry_run=True)
        assert "DRY RUN" in reports[0].message
        assert (tmp_path / "params.h").read_text() == original  # Unchanged


# ============================================================================
# Test Class: Check Command (CI Mode)
# ============================================================================


class TestCheckCommand:
    """Test the check() function for CI/CD usage."""

    def test_in_sync_reports_ok(self, tmp_path):
        """check() returns ALREADY_IN_SYNC when files match."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="both")
        sync(config_path=config)

        reports = check(config_path=config)
        assert reports[0].result == SyncResult.ALREADY_IN_SYNC

    def test_out_of_sync_reports_conflict(self, tmp_path):
        """check() detects when files are out of sync."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="both")
        sync(config_path=config)

        # Modify source
        new_csv = CSV_CONTENT.replace("3000.0", "1234.0")
        (tmp_path / "params.csv").write_text(new_csv)

        reports = check(config_path=config)
        assert reports[0].result == SyncResult.CONFLICT

    def test_missing_target_reports_error(self, tmp_path):
        """check() reports error when target doesn't exist."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.json",
                               direction="both")

        reports = check(config_path=config)
        assert reports[0].result == SyncResult.ERROR


# ============================================================================
# Test Class: Edge Cases
# ============================================================================


class TestEdgeCases:
    """Test edge cases and error handling."""

    def test_missing_source_file(self, tmp_path):
        """Sync reports error when source doesn't exist."""
        config = create_config(tmp_path, "nonexistent.csv", "params.h")
        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.ERROR

    def test_empty_csv(self, tmp_path):
        """Empty CSV (header only) syncs zero constants."""
        (tmp_path / "empty.csv").write_text("Name,Value,Unit,Description\n")
        config = create_config(tmp_path, "empty.csv", "empty.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 0

    def test_single_constant(self, tmp_path):
        """Single constant syncs correctly."""
        (tmp_path / "one.csv").write_text("Name,Value,Unit,Description\nPI,3.14159,,Pi\n")
        config = create_config(tmp_path, "one.csv", "one.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 1
        assert "PI" in (tmp_path / "one.h").read_text()

    def test_large_number_of_constants(self, tmp_path):
        """Many constants sync without issue."""
        lines = ["Name,Value,Unit,Description"]
        for i in range(100):
            lines.append(f"CONST_{i},{i * 1.5},,Constant {i}")
        (tmp_path / "big.csv").write_text("\n".join(lines) + "\n")
        config = create_config(tmp_path, "big.csv", "big.h")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET
        assert reports[0].count == 100

    def test_special_characters_in_description(self, tmp_path):
        """Descriptions with special chars don't break output."""
        csv_content = 'Name,Value,Unit,Description\nRATIO,1.5,,">= 1.0 && <= 2.0"\n'
        (tmp_path / "special.csv").write_text(csv_content)
        config = create_config(tmp_path, "special.csv", "special.json")

        reports = sync(config_path=config)
        assert reports[0].result == SyncResult.SYNCED_SOURCE_TO_TARGET

    def test_backup_created_on_overwrite(self, tmp_path):
        """Backup is created when target is overwritten."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        config = create_config(tmp_path, "params.csv", "params.h")

        # First sync creates target
        sync(config_path=config)

        # Modify source to trigger overwrite
        new_csv = CSV_CONTENT.replace("3000.0", "5000.0")
        (tmp_path / "params.csv").write_text(new_csv)

        # Second sync should backup
        sync(config_path=config)
        backup_dir = tmp_path / ".consync" / "backups"
        assert backup_dir.exists()
        backups = list(backup_dir.iterdir())
        assert len(backups) > 0


# ============================================================================
# Test Class: Format-Specific Rendering
# ============================================================================


class TestFormatRendering:
    """Test format-specific rendering features."""

    def test_c_header_define_style(self, tmp_path):
        """C header renders with #define style."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.h
                direction: source_to_target
                output_style: define
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)
        content = (tmp_path / "params.h").read_text()
        assert "#define" in content

    def test_c_header_const_style(self, tmp_path):
        """C header renders with const style."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.h
                direction: source_to_target
                output_style: const
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)
        content = (tmp_path / "params.h").read_text()
        assert "const" in content

    def test_verilog_module_name(self, tmp_path):
        """Verilog renders with module wrapper."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.v
                direction: source_to_target
                module_name: motor_params
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)
        content = (tmp_path / "params.v").read_text()
        assert "module motor_params" in content
        assert "endmodule" in content

    def test_csharp_namespace(self, tmp_path):
        """C# renders with namespace."""
        (tmp_path / "params.csv").write_text(CSV_CONTENT)
        cfg_text = textwrap.dedent("""\
            mappings:
              - source: params.csv
                target: params.cs
                direction: source_to_target
                namespace: Motor.Config
        """)
        config = tmp_path / ".consync.yaml"
        config.write_text(cfg_text)

        sync(config_path=config)
        content = (tmp_path / "params.cs").read_text()
        assert "Motor.Config" in content
